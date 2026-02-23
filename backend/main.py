from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
import os
import json
import httpx

app = FastAPI(title="Smart Excel Analyzer API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}"

def unmerge_and_fill(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    merged_ranges = list(ws.merged_cells.ranges)
    for merge_range in merged_ranges:
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def detect_header_row(df_raw):
    for i, row in df_raw.iterrows():
        non_null = row.dropna()
        if len(non_null) >= 2:
            str_count = sum(1 for v in non_null if isinstance(v, str))
            if str_count >= len(non_null) * 0.5:
                return i
    return 0

def clean_dataframe(df):
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')
    new_cols = []
    seen = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str.startswith('Unnamed') or col_str == 'nan':
            col_str = f'Colonne_{len(new_cols)+1}'
        if col_str in seen:
            seen[col_str] += 1
            col_str = f'{col_str}_{seen[col_str]}'
        else:
            seen[col_str] = 0
        new_cols.append(col_str)
    df.columns = new_cols
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].ffill()
    for col in df.columns:
        if df[col].dtype == object:
            mask = df[col].astype(str).str.lower().str.contains(
                'total|sous.total|somme|sum|grand total',
                regex=True, na=False
            )
            df = df[~mask]
    df = df.reset_index(drop=True)
    return df

def is_phone_number(series):
    """Détecte si une colonne est un numéro de téléphone"""
    clean = series.dropna()
    if len(clean) == 0:
        return False
    try:
        nums = pd.to_numeric(clean, errors='coerce').dropna()
        if len(nums) == 0:
            return False
        # Téléphones camerounais : 9 chiffres, commencent par 6
        return nums.between(600000000, 699999999).sum() > len(nums) * 0.5
    except:
        return False

def is_matricule(series):
    """Détecte si une colonne est un matricule"""
    clean = series.dropna().astype(str)
    # Matricule : contient des tirets et des chiffres
    return clean.str.contains(r'\d{3}-\d{7}-\d', regex=True).sum() > len(clean) * 0.5

def smart_read_excel(file_bytes):
    try:
        cleaned_buffer = unmerge_and_fill(file_bytes)
    except Exception:
        cleaned_buffer = BytesIO(file_bytes)
    try:
        df_raw = pd.read_excel(cleaned_buffer, header=None)
    except Exception:
        df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
    header_row = detect_header_row(df_raw)
    cleaned_buffer.seek(0)
    try:
        df = pd.read_excel(cleaned_buffer, header=header_row)
    except Exception:
        df = pd.read_excel(BytesIO(file_bytes), header=header_row)
    df = clean_dataframe(df)
    for col in df.columns:
        try:
            converted = pd.to_numeric(
                df[col].astype(str).str.replace(' ', '').str.replace(',', '.'),
                errors='coerce'
            )
            if converted.notna().sum() > len(df) * 0.5:
                df[col] = converted
        except Exception:
            pass
    for col in df.columns:
        if df[col].dtype == object:
            try:
                converted = pd.to_datetime(df[col], infer_datetime_format=True, errors='coerce')
                if converted.notna().sum() > len(df) * 0.5:
                    df[col] = converted
            except Exception:
                pass
    return df

def extract_basic_stats(df):
    """Extraction statistiques de base — calcul local rapide"""
    number_cols = []
    date_cols = []
    category_cols = []
    boolean_cols = []
    ignored_cols = []
    kpis = []
    charts = []
    alerts = []
    anomalies = []

    for col in df.columns:
        series = df[col]

        # Ignorer téléphones et matricules
        if is_phone_number(series):
            ignored_cols.append(col)
            continue
        if is_matricule(series):
            ignored_cols.append(col)
            continue

        if pd.api.types.is_datetime64_any_dtype(series):
            date_cols.append(col)
            continue

        unique_lower = series.dropna().astype(str).str.lower().unique()
        bool_keywords = {"oui","non","yes","no","true","false","présent","absent","actif","inactif"}
        if set(unique_lower).issubset(bool_keywords):
            boolean_cols.append(col)
            continue

        if pd.api.types.is_numeric_dtype(series):
            number_cols.append(col)
            continue

        category_cols.append(col)

    print(f"[STATS] Colonnes numériques: {number_cols}")
    print(f"[STATS] Colonnes ignorées: {ignored_cols}")

    # KPIs
    for col in number_cols:
        clean = df[col].dropna()
        if len(clean) == 0:
            continue
        kpis.append({
            "column": col,
            "total": round(float(clean.sum()), 2),
            "average": round(float(clean.mean()), 2),
            "min": round(float(clean.min()), 2),
            "max": round(float(clean.max()), 2),
            "count": int(clean.count())
        })
        if clean.std() > 0 and clean.max() > clean.mean() + 2 * clean.std():
            alerts.append({
                "type": "warning",
                "message": f"Valeur élevée détectée dans '{col}': max={round(float(clean.max()), 2)}, moyenne={round(float(clean.mean()), 2)}"
            })

    # Graphiques Date + Nombre
    for date_col in date_cols:
        for num_col in number_cols:
            try:
                time_series = df.groupby(df[date_col].dt.to_period("M"))[num_col].sum()
                values = [round(float(v), 2) for v in time_series.values]
                labels = [str(p) for p in time_series.index]
                if len(values) >= 2 and values[-1] < values[-2] * 0.8:
                    alerts.append({
                        "type": "danger",
                        "message": f"Baisse de plus de 20% sur '{num_col}'"
                    })
                charts.append({
                    "type": "line",
                    "title": f"Évolution de {num_col} par mois",
                    "labels": labels,
                    "values": values,
                    "x_col": date_col,
                    "y_col": num_col
                })
            except Exception:
                pass

    # Graphiques Catégorie + Nombre — limité aux colonnes utiles
    for cat_col in category_cols:
        # Ignorer les colonnes avec trop de valeurs uniques (matricules, noms)
        if df[cat_col].nunique() > 20:
            continue
        for num_col in number_cols:
            try:
                grouped = df.groupby(cat_col)[num_col].sum().sort_values(ascending=False).head(10)
                if len(grouped) > 1:
                    charts.append({
                        "type": "bar",
                        "title": f"{num_col} par {cat_col}",
                        "labels": [str(l) for l in grouped.index.tolist()],
                        "values": [round(float(v), 2) for v in grouped.values.tolist()],
                        "x_col": cat_col,
                        "y_col": num_col
                    })
            except Exception:
                pass

    # Graphiques Booléen
    for bool_col in boolean_cols:
        try:
            counts = df[bool_col].astype(str).str.lower().value_counts()
            charts.append({
                "type": "donut",
                "title": f"Répartition de {bool_col}",
                "labels": counts.index.tolist(),
                "values": [int(v) for v in counts.values.tolist()]
            })
        except Exception:
            pass

    # Anomalies
    for col in number_cols:
        try:
            clean = df[col].dropna()
            if len(clean) < 3:
                continue
            mean = clean.mean()
            std = clean.std()
            if std == 0:
                continue
            outliers = df[abs(df[col] - mean) > 3 * std]
            if not outliers.empty:
                anomalies.append({
                    "column": col,
                    "count": len(outliers),
                    "message": f"{len(outliers)} valeur(s) aberrante(s) dans '{col}'"
                })
        except Exception:
            pass

    return {
        "kpis": kpis,
        "charts": charts,
        "alerts": alerts,
        "anomalies": anomalies,
        "number_cols": number_cols,
        "date_cols": date_cols,
        "category_cols": category_cols,
        "boolean_cols": boolean_cols,
        "ignored_cols": ignored_cols
    }

async def gemini_full_analysis(df, basic_stats):
    """Gemini fait l'analyse complète et intelligente"""
    try:
        apercu = df.head(20).to_string()
        stats_summary = {
            "colonnes_numeriques": basic_stats["number_cols"],
            "colonnes_dates": basic_stats["date_cols"],
            "colonnes_categories": basic_stats["category_cols"],
            "colonnes_ignorees": basic_stats["ignored_cols"],
            "kpis": basic_stats["kpis"],
            "alertes_detectees": basic_stats["alerts"],
            "anomalies": basic_stats["anomalies"],
        }

        prompt = f"""Tu es un expert analyste de données senior pour PME africaines.
Analyse ce fichier Excel de façon complète, intelligente et professionnelle.

APERÇU DES DONNÉES (20 premières lignes) :
{apercu}

STATISTIQUES CALCULÉES :
{json.dumps(stats_summary, ensure_ascii=False, default=str)}

INSTRUCTIONS IMPORTANTES :
- Identifie précisément le domaine (RH, finance, sport, comptabilité, inventaire, etc.)
- Les colonnes ignorées sont des téléphones ou matricules — ne les analyse pas comme des chiffres
- Analyse en profondeur ce que représentent vraiment ces données
- Génère des insights SPÉCIFIQUES aux données — pas de messages génériques
- Donne des conseils CONCRETS et ACTIONNABLES adaptés au contexte camerounais
- Identifie des patterns cachés que l'oeil humain ne voit pas facilement
- Si c'est une fiche RH/CNPS — analyse les charges sociales, CV (congés?), CS (charges sociales?)
- Le plan d'action doit être réaliste et immédiatement applicable

Réponds UNIQUEMENT en JSON valide avec cette structure exacte :
{{
  "domaine": "string",
  "contexte": "string (description précise de ce que contient ce fichier)",
  "resume_executif": "string (3-4 phrases percutantes sur la situation globale)",
  "score_sante": number (0-100),
  "score_explication": "string (pourquoi ce score)",
  "insights": [
    {{
      "titre": "string",
      "observation": "string (fait précis tiré des données)",
      "conseil": "string (action concrète et immédiate)",
      "priorite": "haute|moyenne|faible",
      "icone": "emoji"
    }}
  ],
  "points_forts": ["string", "string", "string"],
  "points_faibles": ["string", "string", "string"],
  "opportunites": ["string", "string"],
  "risques": ["string", "string"],
  "plan_action": [
    {{
      "priorite": number,
      "action": "string",
      "delai": "string",
      "responsable": "string",
      "impact": "string"
    }}
  ],
  "conclusion": "string (message final motivant)"
}}"""

        print(f"[GEMINI] Envoi requête...")
        async with httpx.AsyncClient(timeout=45.0) as client:
            response = await client.post(
                GEMINI_URL,
                json={
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {
                        "temperature": 0.2,
                        "maxOutputTokens": 4096
                    }
                }
            )
            print(f"[GEMINI] Status: {response.status_code}")
            data = response.json()
            print(f"[GEMINI] Response keys: {list(data.keys())}")

            if "error" in data:
                print(f"[GEMINI] ERREUR API: {data['error']}")
                raise Exception(f"Gemini error: {data['error']}")

            text = data["candidates"][0]["content"]["parts"][0]["text"]
            print(f"[GEMINI] Texte reçu ({len(text)} chars)")

            text = text.strip()
            if text.startswith("```json"):
                text = text[7:]
            if text.startswith("```"):
                text = text[3:]
            if text.endswith("```"):
                text = text[:-3]
            text = text.strip()

            result = json.loads(text)
            print(f"[GEMINI] JSON parsé avec succès - domaine: {result.get('domaine')}")
            return result

    except Exception as e:
        print(f"[GEMINI] EXCEPTION: {type(e).__name__}: {str(e)}")
        return {
            "domaine": "Données générales",
            "contexte": "Fichier Excel analysé",
            "resume_executif": f"Erreur analyse IA: {str(e)}",
            "score_sante": 75,
            "score_explication": "Score par défaut",
            "insights": [],
            "points_forts": [],
            "points_faibles": [],
            "opportunites": [],
            "risques": [],
            "plan_action": [],
            "conclusion": "Continuez votre excellent travail."
        }

async def gemini_compare(data1, data2, file1_name, file2_name):
    """Gemini compare deux fichiers de façon intelligente"""
    try:
        prompt = f"""Tu es un expert analyste de données senior.
Compare ces deux fichiers Excel de façon intelligente et professionnelle.

FICHIER 1 : {file1_name}
{json.dumps(data1, ensure_ascii=False, default=str)}

FICHIER 2 : {file2_name}
{json.dumps(data2, ensure_ascii=False, default=str)}

Analyse les différences, tendances et évolutions entre les deux fichiers.

Réponds UNIQUEMENT en JSON valide :
{{
  "resume_comparaison": "string (3-4 phrases sur les différences principales)",
  "evolution_globale": "positive|negative|stable",
  "score_file1": number,
  "score_file2": number,
  "differences_cles": [
    {{
      "indicateur": "string",
      "valeur_file1": "string",
      "valeur_file2": "string",
      "evolution": "string",
      "interpretation": "string"
    }}
  ],
  "points_amelioration": ["string", "string", "string"],
  "points_regression": ["string", "string"],
  "recommandations": ["string", "string", "string"],
  "conclusion": "string"
}}"""

        async with httpx.AsyncClient(timeout=45.0) as client:
            response = await client.post(
                GEMINI_URL,
                json={
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {
                        "temperature": 0.2,
                        "maxOutputTokens": 2048
                    }
                }
            )
            data = response.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"]
            text = text.strip()
            if text.startswith("```json"):
                text = text[7:]
            if text.startswith("```"):
                text = text[3:]
            if text.endswith("```"):
                text = text[:-3]
            return json.loads(text.strip())

    except Exception as e:
        print(f"[GEMINI COMPARE] EXCEPTION: {str(e)}")
        return {
            "resume_comparaison": "Comparaison effectuée.",
            "evolution_globale": "stable",
            "score_file1": 70,
            "score_file2": 70,
            "differences_cles": [],
            "points_amelioration": [],
            "points_regression": [],
            "recommandations": [],
            "conclusion": ""
        }

@app.get("/")
def root():
    return {"message": "Smart Excel Analyzer API is running 🚀"}

@app.post("/analyze")
async def analyze(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = smart_read_excel(contents)

        print(f"[ANALYZE] Fichier lu: {len(df)} lignes, {len(df.columns)} colonnes")
        print(f"[ANALYZE] Colonnes: {list(df.columns)}")

        basic_stats = extract_basic_stats(df)
        ai_insights = await gemini_full_analysis(df, basic_stats)

        result = {
            "summary": {
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "missing_values": int(df.isnull().sum().sum()),
                "columns_list": list(df.columns)
            },
            "kpis": basic_stats["kpis"],
            "charts": basic_stats["charts"],
            "alerts": basic_stats["alerts"],
            "anomalies": basic_stats["anomalies"],
            "ai_insights": ai_insights
        }

        return {"status": "success", "data": result}
    except Exception as e:
        print(f"[ANALYZE] ERREUR: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.post("/compare")
async def compare(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    try:
        contents1 = await file1.read()
        contents2 = await file2.read()

        df1 = smart_read_excel(contents1)
        df2 = smart_read_excel(contents2)

        stats1 = extract_basic_stats(df1)
        stats2 = extract_basic_stats(df2)

        ai_compare = await gemini_compare(
            {"kpis": stats1["kpis"], "summary": {"rows": len(df1), "cols": len(df1.columns)}},
            {"kpis": stats2["kpis"], "summary": {"rows": len(df2), "cols": len(df2.columns)}},
            file1.filename,
            file2.filename
        )

        return {
            "status": "success",
            "data": {
                "file1": file1.filename,
                "file2": file2.filename,
                "stats1": stats1,
                "stats2": stats2,
                "ai_compare": ai_compare
            }
        }
    except Exception as e:
        print(f"[COMPARE] ERREUR: {str(e)}")
        return {"status": "error", "message": str(e)}
